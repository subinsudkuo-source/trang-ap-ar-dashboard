from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parent
TARGET_PERIOD_TH = "เมษายน 2569"

AP_FILES = {
    "รพ.ตรัง": "trang.xlsx",
    "รพ.กันตัง": "kantang.xlsx",
    "รพ.ย่านตาขาว": "yan_ta_khao.xlsx",
    "รพ.ปะเหลียน": "palian.xlsx",
    "รพ.สิเกา": "sikao.xlsx",
    "รพ.ห้วยยอด": "huai_yot.xlsx",
    "รพ.วังวิเศษ": "wang_wiset.xlsx",
    "รพ.นาโยง": "na_yong.xlsx",
    "รพ.รัษฎา": "ratsada.xlsx",
    "รพ.หาดสำราญ": "hat_samran_openpyxl.xlsx",
}

HOSPITALS = list(AP_FILES.keys())

ACCOUNT_CODES = {
    "AR_OP_UC_OUT_CUP_IN_PROVINCE": "1102050101.203",
    "AP_OP_UC_OUT_CUP_IN_PROVINCE": "2101020199.202",
}


def clean_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_hospital(value: str) -> str:
    text = clean_text(value)
    text = re.sub(r"^\d+\s*", "", text).strip()
    if "," in text:
        text = text.split(",", 1)[0].strip()
    text = text.replace("โรงพยาบาล", "รพ.")
    text = text.replace("รพ_", "รพ.")
    text = text.replace("รพศ_", "รพ.")
    text = text.replace("รพช_", "รพ.")
    text = text.replace(",", "").strip()
    text = text.replace("รพ. ", "รพ.")
    text = text.replace("เฉลิมพระเกียรติ 80 พรรษา", "").strip()
    text = text.replace("๘๐ พรรษา", "").strip()
    text = text.replace("80 พรรษา", "").strip()
    text = text.replace("ฯ", "").strip()
    text = re.sub(r"\s+", " ", text)
    if text and not text.startswith("รพ."):
        text = "รพ." + text
    if "หาดสำราญ" in text:
        return "รพ.หาดสำราญ"
    return text


def number(value):
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    if not text:
        return 0.0
    return float(text)


def parse_ap_workbook(owner: str, filename: str):
    path = ROOT / filename
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [clean_text(c.value) for c in ws[4]]
    rows = []
    if not (header and "ลำดับ" in header[0]):
        for source_row in ws.iter_rows(values_only=True):
            creditor = normalize_hospital(source_row[0] if source_row else "")
            if creditor not in HOSPITALS:
                continue
            total = number(source_row[10] if len(source_row) > 10 else None)
            record = {
                "period": TARGET_PERIOD_TH,
                "payer_hospital": owner,
                "creditor_hospital": creditor,
                "source_file": filename,
                "amount_total": total,
                "รวมเป็นเงิน": total,
            }
            rows.append(record)
        return rows

    total_idx = next(
        (idx for idx, label in enumerate(header) if "รวมเป็นเงิน" in label or label == "รวม"),
        min(len(header), 19) - 1,
    )
    for row in ws.iter_rows(min_row=5, values_only=True):
        creditor_label = clean_text(row[1] if len(row) > 1 else "")
        if not creditor_label:
            continue
        if "รวมเจ้าหนี้" in creditor_label or "หมายเหตุ" in creditor_label:
            break
        creditor = normalize_hospital(creditor_label)
        if creditor not in HOSPITALS:
            continue
        record = {
            "period": TARGET_PERIOD_TH,
            "payer_hospital": owner,
            "creditor_hospital": creditor,
            "source_file": filename,
        }
        for idx, label in enumerate(header[: total_idx + 1]):
            if not label or idx in (0, 1):
                continue
            key = label.strip()
            record[key] = number(row[idx] if idx < len(row) else None)
        record["amount_total"] = number(row[total_idx] if total_idx < len(row) else None)
        record["รวมเป็นเงิน"] = record["amount_total"]
        rows.append(record)
    return rows


def parse_trial_balance():
    path = ROOT / "trial_balance_apr69.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    max_col = ws.max_column
    blocks = []
    for start in range(1, max_col + 1, 12):
        title = clean_text(rows[0][start - 1] if start - 1 < len(rows[0]) else None)
        if not title:
            continue
        hospitals = []
        for col in range(start + 2, min(start + 12, max_col + 1)):
            hospitals.append(normalize_hospital(rows[2][col - 1] if col - 1 < len(rows[2]) else None))
        blocks.append({"start": start, "title": title, "hospitals": hospitals})

    extracted = []
    for block in blocks:
        start = block["start"]
        period = block["title"].replace("งบการเงิน (งบทดลอง) สิ้น", "").strip()
        for code_name, code in ACCOUNT_CODES.items():
            for source_row in rows:
                if clean_text(source_row[start - 1] if start - 1 < len(source_row) else None) != code:
                    continue
                account_name = clean_text(source_row[start] if start < len(source_row) else None)
                for offset, hospital in enumerate(block["hospitals"], start=2):
                    if hospital in HOSPITALS:
                        extracted.append(
                            {
                                "period": period,
                                "account_key": code_name,
                                "account_code": code,
                                "account_name": account_name,
                                "hospital": hospital,
                                "amount": number(
                                    source_row[start + offset - 1]
                                    if start + offset - 1 < len(source_row)
                                    else None
                                ),
                            }
                        )
                break
    return extracted


def main():
    ledger_rows = []
    for owner, filename in AP_FILES.items():
        ledger_rows.extend(parse_ap_workbook(owner, filename))

    matrix = {
        payer: {creditor: 0.0 for creditor in HOSPITALS}
        for payer in HOSPITALS
    }
    for row in ledger_rows:
        matrix[row["payer_hospital"]][row["creditor_hospital"]] = row["amount_total"]

    trial_rows = parse_trial_balance()
    target_trial = [r for r in trial_rows if TARGET_PERIOD_TH in r["period"]]
    tb_by_account_hospital = {
        (r["account_key"], r["hospital"]): r["amount"]
        for r in target_trial
    }

    reconciliation = []
    for hospital in HOSPITALS:
        ap_ledger_total = sum(matrix[hospital].values())
        ar_from_counterparties = sum(matrix[payer][hospital] for payer in HOSPITALS)
        ap_tb = tb_by_account_hospital.get(("AP_OP_UC_OUT_CUP_IN_PROVINCE", hospital), 0.0)
        ar_tb = tb_by_account_hospital.get(("AR_OP_UC_OUT_CUP_IN_PROVINCE", hospital), 0.0)
        reconciliation.append(
            {
                "hospital": hospital,
                "ap_ledger_total": ap_ledger_total,
                "ap_trial_balance": ap_tb,
                "ap_difference": ap_ledger_total - ap_tb,
                "ar_from_counterparties": ar_from_counterparties,
                "ar_trial_balance": ar_tb,
                "ar_difference": ar_from_counterparties - ar_tb,
            }
        )

    trang_rows = []
    for hospital in HOSPITALS:
        if hospital == "รพ.ตรัง":
            continue
        trang_payable = matrix["รพ.ตรัง"][hospital]
        hospital_payable_to_trang = matrix[hospital]["รพ.ตรัง"]
        trang_rows.append(
            {
                "community_hospital": hospital,
                "trang_payable_to_community": trang_payable,
                "community_payable_to_trang": hospital_payable_to_trang,
                "net_for_trang": hospital_payable_to_trang - trang_payable,
            }
        )

    output = {
        "period": TARGET_PERIOD_TH,
        "hospitals": HOSPITALS,
        "ledger_rows": ledger_rows,
        "matrix": matrix,
        "trial_balance_rows": trial_rows,
        "trial_balance_target_rows": target_trial,
        "reconciliation": reconciliation,
        "trang_comparison": trang_rows,
    }
    out_path = ROOT / "dashboard_data.json"
    out_path.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(
        {
            "ledger_rows": len(ledger_rows),
            "trial_rows": len(trial_rows),
            "target_trial_rows": len(target_trial),
            "ap_total": sum(r["amount_total"] for r in ledger_rows),
            "output": str(out_path),
        },
        ensure_ascii=False,
        indent=2,
    ))


if __name__ == "__main__":
    main()
